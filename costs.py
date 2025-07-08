import boto3
import pandas as pd
from datetime import datetime
from dateutil.relativedelta import relativedelta
from openpyxl import load_workbook

session = boto3.Session(profile_name='root')
ce = session.client('ce')
organizations = session.client('organizations')

COST_TYPE = 'AmortizedCost'


def get_dates(month: int, year: int):
    """
    Returns the start and end date of the previous month and the next month
    
    Parameters:
        month (int): The month to be used as reference
        year (int): The year to be used as reference
    
    Returns:
        start_date (str): The start date of the previous month in the format 'YYYY-MM-DD'
        end_date (str): The end date of the next month in the format 'YYYY-MM-DD'
    """
    start_date = datetime(year, month, 1) - relativedelta(months=1)
    end_date = datetime(year, month, 1) + relativedelta(months=1)
    
    start_date = start_date.strftime("%Y-%m-%d")
    end_date = end_date.strftime("%Y-%m-%d")
    
    return start_date, end_date

def all_account_costs(month: int, year: int):
    """
    Returns a DataFrame with the costs for each account, from the past and current month
    
    Parameters:
        month (int): The month to be used as reference
        year (int): The year to be used as reference
        
    Returns:
        df (DataFrame): A DataFrame with the following columns:
            - "Account ID": The ID of the account
            - "Account Name": The name of the account
            - "Past Month": The total cost of the account in the past month
            - "Current Month": The total cost of the account in the current month
            - "Absolute Diff": The absolute difference between the costs of the past and current month
            - "Relative Diff (%)": The relative difference between the costs of the past and current month
            - "Details": An empty column to be filled
    """
    start_date, end_date = get_dates(month, year)

    account_list = organizations.list_accounts()
    data = []

    for account in account_list['Accounts']:
        account_id = account['Id']
        account_name = account['Name']
        response = ce.get_cost_and_usage(
            TimePeriod={
                'Start': start_date,
                'End': end_date
            },
            Granularity='MONTHLY',
            Metrics=[COST_TYPE],
            Filter={
                'Dimensions': {
                    'Key': 'LINKED_ACCOUNT',
                    'Values': [
                        account_id
                    ]
                }
            }
        )

        amount1 = float(response['ResultsByTime'][0]['Total'][COST_TYPE]['Amount'])
        amount2 = float(response['ResultsByTime'][1]['Total'][COST_TYPE]['Amount'])

        absolute_change = amount2 - amount1
        percent_change = (absolute_change / amount1) * 100 if amount1 != 0 else 0

        data.append([account_id, account_name, amount1, amount2, absolute_change, percent_change])

    # Create a DataFrame
    df = pd.DataFrame(data, columns=["Account ID", "Account Name", "Past Month", "Current Month", "Absolute Diff", "Relative Diff (%)"])
    # Add a column "Details" to be filled out
    df["Details"] = None
    # Sort the DataFrame from highest to lowest absolute (|x|) value of the "Absolute Diff" column
    df = df.reindex(df['Absolute Diff'].abs().sort_values(ascending=False).index)
    
    return df

def account_services_costs(month: int, year:int, account_id):
    """
    Returns a DataFrame with the costs for each service of a specific account, from the past and current month
    
    Parameters:
        month (int): The month to be used as reference
        year (int): The year to be used as reference
        account_id (str): The ID of the account
        
    Returns:
        df (DataFrame): A DataFrame with the following columns:
            - "Service": The name of the service
            - "Past Month": The total cost of the service in the past month
            - "Current Month": The total cost of the service in the current month
            - "Absolute Diff": The absolute difference between the costs of the past and current month
            - "Relative Diff (%)": The relative difference between the costs of the past and current month
            - "Details": An empty column to be filled
    """
    start_date, end_date = get_dates(month, year)

    print(f"Getting costs for account {account_id} from {start_date} to {end_date}")
    
    response = ce.get_cost_and_usage(
        TimePeriod={
            'Start': start_date,
            'End': end_date
        },
        Granularity='MONTHLY',
        GroupBy=[
            {
                'Type': 'DIMENSION',
                'Key': 'SERVICE'
            }
        ],
        Metrics=[COST_TYPE],
        Filter={
            'Dimensions': {
                'Key': 'LINKED_ACCOUNT',
                'Values': [
                    account_id
                ]
            }
        }
    )

    data = []
    
    # create a dataframe with the following columns ["Service", "Cost_TimePeriod_0", "Cost_TimePeriod_1", "Cost_TimePeriod_x", "Absolute Diff", "Relative Diff (%)"]
    for group in response['ResultsByTime'][0]['Groups']:
        service = group['Keys'][0]
        amount1 = float(group['Metrics'][COST_TYPE]['Amount'])
        
        amount2 = 0
        for group2 in response['ResultsByTime'][1]['Groups']:
            if group2['Keys'][0] == service:
                amount2 = float(group2['Metrics'][COST_TYPE]['Amount'])
                break
        
        amount_sum = amount1 + amount2
        absolute_change = amount2 - amount1
        percent_change = (absolute_change / amount1) * 100 if amount1 != 0 else 0

        data.append([service, amount_sum, amount1, amount2, absolute_change, percent_change])
            
    df = pd.DataFrame(data, columns=["Service", "Sum", "Past Month", "Current Month", "Absolute Diff", "Relative Diff (%)"])
    
    # Format the DataFrame to display numbers with two decimal places
    pd.options.display.float_format = '{:.2f}'.format
    # Add a column "Details" to be filled out
    df["Details"] = None
    # Sort the DataFrame from highest to lowest absolute (|x|) value of the "Absolute Diff" column
    df = df.reindex(df['Absolute Diff'].abs().sort_values(ascending=False).index)
    # show only the first 10 rows
    df = df.head(10)
    return df

def service_usage_type_costs(month: int, year: int, account_id, service):
    """
    Returns a DataFrame with the costs for each usage type of a specific service of a specific account, from the past and current month
    
    Parameters:
        month (int): The month to be used as reference
        year (int): The year to be used as reference
        account_id (str): The ID of the account
        service (str): The name of the service
        
    Returns:
        df (DataFrame): A DataFrame with the following columns:
            - "Usage Type": The name of the usage type
            - "Past Month": The total cost of the usage type in the past month
            - "Current Month": The total cost of the usage type in the current month
            - "Absolute Diff": The absolute difference between the costs of the past and current month
            - "Relative Diff (%)": The relative difference between the costs of the past and current month
            - "Details": An empty column to be filled
    """
    start_date, end_date = get_dates(month, year)

    response = ce.get_cost_and_usage(
        TimePeriod={
            'Start': start_date,
            'End': end_date
        },
        Granularity='MONTHLY',
        GroupBy=[
            {
                'Type': 'DIMENSION',
                'Key': 'USAGE_TYPE'
            }
        ],
        Metrics=[COST_TYPE],
        Filter={
            'And': [
                {
                    'Dimensions': {
                        'Key': 'LINKED_ACCOUNT',
                        'Values': [
                            account_id
                        ]
                    }
                },
                {
                    'Dimensions': {
                        'Key': 'SERVICE',
                        'Values': [
                            service
                        ]
                    }
                }
            ]
        }
    )

    data = []
    
    # print response to a json file
    with open('response.json', 'w') as f:
        import json
        json.dump(response, f, indent=4)
    
    # Create a dictionary to store data from both time periods
    usage_data = {}

    # Process the first time period
    for group in response['ResultsByTime'][0]['Groups']:
        usage_type = group['Keys'][0]
        amount1 = float(group['Metrics'][COST_TYPE]['Amount'])
        usage_data[usage_type] = {'amount1': amount1, 'amount2': 0}

    # Process the second time period
    for group in response['ResultsByTime'][1]['Groups']:
        usage_type = group['Keys'][0]
        amount2 = float(group['Metrics'][COST_TYPE]['Amount'])
        if usage_type in usage_data:
            usage_data[usage_type]['amount2'] = amount2
        else:
            usage_data[usage_type] = {'amount1': 0, 'amount2': amount2}

    # Create the final dataframe
    for usage_type, amounts in usage_data.items():
        amount1 = amounts['amount1']
        amount2 = amounts['amount2']
        amount_sum = amount1 + amount2
        absolute_change = amount2 - amount1
        percent_change = (absolute_change / amount1) * 100 if amount1 != 0 else 0

        data.append([usage_type, amount_sum, amount1, amount2, absolute_change, percent_change])
            
    df = pd.DataFrame(data, columns=["Usage Type", "Sum", "Past Month", "Current Month", "Absolute Diff", "Relative Diff (%)"])
    
    # Format the DataFrame to display numbers with two decimal places
    pd.options.display.float_format = '{:.2f}'.format
    # Add a column "Details" to be filled out
    df["Details"] = None
    # Sort the DataFrame from highest to lowest absolute (|x|) value of the "Absolute Diff" column
    df = df.reindex(df['Absolute Diff'].abs().sort_values(ascending=False).index)
    # show only the first 10 rows
    df = df.head(10)
    return df

def clean_excel(file_name: str):
    """
    Adjusts the column widths of an Excel file
    
    Parameters:
        file_name (str): The name of the Excel file to be adjusted
    """ 
    workbook = load_workbook(file_name)

    for sheet in workbook.sheetnames:
        worksheet = workbook[sheet]
        # Adjust the column widths
        for col in worksheet.columns:
            max_length = 0
            column = col[0].column_letter  # Get the column name
            for cell in col:
                try:
                    cell_value = str(cell.value)
                    if len(cell_value) > max_length:
                        max_length = len(cell_value)
                except:
                    pass
            adjusted_width = (max_length + 2)
            worksheet.column_dimensions[column].width = adjusted_width

    # Save the workbook
    workbook.save(file_name)


month, year = 6, 2025
excel_name = "output.xlsx"

with pd.ExcelWriter(excel_name) as writer:
    df_all_accounts = all_account_costs(month, year)
    df_all_accounts.to_excel(writer, sheet_name='Accounts', index=False, float_format="%.2f")
    account_list = organizations.list_accounts()
    # For each account, get the costs of the services and usage types
    for account in account_list['Accounts']:
        print(f"Processing account: {account['Name']}, {account['Id']}")
        
        df = account_services_costs(month, year, account['Id'])
        df.to_excel(writer, sheet_name=account['Name'], index=False, float_format="%.2f")
        # add empty line to separate the tables
        pd.DataFrame().to_excel(writer, sheet_name=account['Name'], index=False)
        startrow = df.shape[0] + 2
        
        for service in df['Service']:
            # Create a Dataframe with the Service Name to add on top of the table below
            df_service_name = pd.DataFrame([service], columns=["Service"])
            df_service_name.to_excel(writer, sheet_name=account['Name'], startrow=startrow, index=False, header=False)            
            startrow += 1
            # Create a Dataframe with the Service Usage Type Costs
            df_service = service_usage_type_costs(month, year, account['Id'], service)
            df_service.to_excel(writer, sheet_name=account['Name'], startrow=startrow, index=False, float_format="%.2f")
            startrow += df_service.shape[0] + 2
            # add empty line to separate the tables
            pd.DataFrame().to_excel(writer, sheet_name=account['Name'], startrow=startrow, index=False)
            
clean_excel(excel_name)